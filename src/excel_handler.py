import pandas as pd
import os
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

excel_file = "balance.xlsx"
file_path = f"C:/Users/Usuario/Desktop/Finance App Project/wealth_management/files/{excel_file}"


def read_excel():
    #Llama a checkear que exista el archivo antes de abrirlo. Si no es así
    # lo crea y ahí si va a poder leerlo y retornarlo.
    check_and_create_file(file_path)
    return pd.read_excel(file_path)

def load_titles():
    #Cargo el excel
    book = load_workbook(file_path)
    
    
    #Acá deberíamos tener una variable quizá...
    sheet = book['Gastos']
    
    titles = ['Descripción', 'Monto', 'Categoría', 'Nivel de necesidad']
    
    start_row = 1
    start_col = 'A'
    
    col_idx = column_index_from_string(start_col)
    
    for index, val in enumerate(titles, start=start_row):
        sheet.cell(row=start_row, column=index, value=val)
    book.save(file_path)    

def initialize_excel(df, file_path):
    df1 = pd.DataFrame()
    df2 = pd.DataFrame()
    try:
        # Crea un ExcelWriter
        with pd.ExcelWriter(file_path) as writer:
            # Escribe cada DataFrame en una hoja diferente
            df.to_excel(writer, sheet_name='Gastos', index=False)
            df1.to_excel(writer, sheet_name='Ahorros e inversiones', index=False)
            df2.to_excel(writer, sheet_name='Objetivos', index=False)
        
    except Exception as e:
        print(e)
        
    load_titles() #Cargo los títulos en el Excel
    
    
    
def check_and_create_file(file_path):
    if not os.path.exists(file_path):
        df = pd.DataFrame()        
        initialize_excel(df, file_path)
        print(f"Archivo creado en {file_path}")
        return df
    else:
        print(f"El archivo ya existe en {file_path}")
        return 0
    
def write_excel(data, sheet):
    # Supongamos que tienes un DataFrame con los datos que deseas añadir

    # Cargar el archivo Excel existente
    book = load_workbook(file_path)
    
    sheet = book[sheet]
        
    # Escribir los datos en una columna específica
    start_row = 2  # Ajusta el número de fila inicial si es necesario
    start_col = 'A'  # Ajusta la columna de inicio si es necesario
    
    col_idx = column_index_from_string(start_col)
    
    for idx, val in enumerate(data["Gastos"], start=1):
        sheet.cell(row=start_row, column=idx, value=val)
        
    # Guardar los cambios
    book.save(file_path)

#=============== Variables fundamentales =============== 

# sheets = {
#     "Gastos": ,
#     "Ahorros e inversiones" : ,
#     "Objetivos" : ,
#     }    
    
    
    